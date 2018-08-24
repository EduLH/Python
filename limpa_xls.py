from openpyxl import load_workbook
import pyexcel as pe

loop = 0
ticket = 0
while int(loop) == 0:
    print("informe qual operação deseja realizar:")
    print("1 - Limpeza de planilha do Legacy 500")
    print("2 - Limpeza de planilha do Legacy 650")
    print("3 - Limpeza de planilha do Phenom 100")
    print("4 - Limpeza de planilha do Phenom 300")
    print("5 - Limpeza de planilha do Lineage")
    a = input()
    if int(a) == 1:
        programa = " Legacy 500"
        file_xls = "/home/eduardo/PA_Leg500.xls"
        file_xlsx = "/home/eduardo/PA_Leg500.xlsx"
    elif int(a) == 2:
        programa = " Legacy 650"
        file_xls = "/home/eduardo/PA_Leg650.xls"
        file_xlsx = "/home/eduardo/PA_Leg650.xlsx"
    elif int(a) == 3:
        programa = " Phenom 100"
        file_xls = "/home/eduardo/PA_Ph100.xls"
        file_xlsx = "/home/eduardo/PA_Ph100.xlsx"
    elif int(a) == 4:
        programa = " Phenom 300"
        file_xls = "/home/eduardo/PA_Ph300.xls"
        file_xlsx = "/home/eduardo/PA_Ph300.xlsx"
    elif int(a) == 5:
        programa = " Lineage"
        file_xls = "/home/eduardo/PA_Lineage1000.xls"
        file_xlsx = "/home/eduardo/PA_Lineage1000.xlsx"
    else:
        print("opção não reconhecida")

    print("Você executará limpeza da planilha do" + programa)
    print("Deseja REALMENTE executar esta aca? ( S / N )" )
    check = input()
    if check in ['s', 'S']:
        ticket = 1
    elif check in ['n', 'N']:
        ticket = 0
    else:
        print("Opcao nao reconhecida")

    if ticket == 1:
        pe.save_book_as(file_name=file_xls, dest_file_name=file_xlsx)
        wb = load_workbook(filename=file_xlsx)
        for i in range(0,2):
            sheet = wb.worksheets[i]
            row_count = sheet.max_row
            column_count = sheet.max_column

            for line in range(1, row_count+1):
                for column1 in range(1, column_count+1):
                    if line == 1:
                        conteudo = sheet.cell(row=line, column=column1).value
                        conteudo = str(conteudo).replace(" ", "")
                        sheet.cell(row=line, column=column1).value = conteudo

                    conteudo = sheet.cell(row=line, column=column1).value
                    if '.' in str(conteudo):
                        conteudo = str(conteudo).replace(".", "_")
                        sheet.cell(row=line, column=column1).value = conteudo

        print("O documento do" + programa + " foi limpo")
        wb.save(file_xlsx)
        pe.save_book_as(file_name=file_xlsx, dest_file_name=file_xls)

    ticket = 0
    print("deseja fazer mais alguma operacao?")
    print("0 - sim; 1 - nao")
    loop = input()
print("Tchau :)")
